import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename, data_only=True)
    sheet = wb['transactions']
    for row in range(2, sheet.max_row + 1):
        column = 4
        sheet.cell(row, column).value *= 0.9
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=6, max_col=6)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'G1')
    wb.save(filename)


process_workbook()
